import sqlite3 as sq
from vkbottle.bot import Message
from vkbottle import Bot, Keyboard, KeyboardButtonColor, Text, Callback, GroupEventType, GroupTypes
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from config import *
import mimetypes
import calendar
import requests
from art import tprint
import datetime
import openpyxl
import smtplib
import shutil
import math
import time
import os

#TODO: РАЗДЕЛИТЬ БОТА НА ФАЙЛЫ
def generate():
    next_answer = math.floor(time.time()) + (60 * 60 * 24)
    return next_answer


def create_list_payers(class_name):
    x = 13
    payers_list = []
    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb['Лист1']

    while sheet[f'b{x}'].value.lower() != 'итого':
        payers_list.append((sheet[f'b{x}']).value)
        x += 1
    wb.close()
    return payers_list
    pass


def filling_template(values_dict, class_name):
    x = 13
    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb['Лист1']
    week_date = datetime.datetime.today().weekday()
    if week_date == 0:
        sheet['f12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=6).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()
    elif week_date == 1:
        sheet['g12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=7).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()
    elif week_date == 2:
        sheet['h12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=8).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()
    elif week_date == 3:
        sheet['i12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=9).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()
    elif week_date == 4:
        sheet['j12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=10).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()
    elif week_date == 5:
        sheet['k12'] = datetime.datetime.now().strftime('%d')
        for value in values_dict:
            sheet.cell(row=x, column=11).value = value
            x += 1
            wb.save(f'attachments/Template_{class_name}.xlsx')
            wb.close()


def create_template(doc):
    x = 13
    columns = ['f', 'g', 'h', 'i', 'j', 'k']  # Столбцы, в которых записываются данные об отсутствующих

    try:

        wb = openpyxl.load_workbook(filename=f'{doc}')
        sheet = wb['Лист1']

        sheet['F7'] = None
        sheet['H7'] = None
        sheet['K6'] = None
        sheet['L9'] = None
        sheet['K9'] = '=F12'

        while sheet[f'b{x}'].value != 'итого':
            x += 1
        sheet_offset = sheet['b13'].offset(row=x - 13, column=0)

        for i in columns:  # Запись формул по сложению в строку
            formula = f'=SUM({i}{13}:{i}{sheet_offset.row - 1})'
            sheet[f'{i}{sheet_offset.row}'] = formula

        for i in range(sheet_offset.row - 1, 12, -1):  # Запись формул по сложению в столбец
            formula2 = f'=SUM(f{i}:k{i})'
            sheet[f'l{i}'] = formula2

        formula3 = '=IF('  # Запись формулы по проверке и вычислении итого
        for i in columns:
            formula3 += f'+{i}{sheet_offset.row}'
        formula3 += '='
        for j in range(sheet_offset.row - 1, 12, -1):
            formula3 += f'+l{j}'
        formula3 += ','
        for i in columns:
            formula3 += f'+{i}{sheet_offset.row}'
        formula3 += ',"ошибка")'
        sheet[f'l{sheet_offset.row}'] = formula3

        for i in columns:  # очистка ячеек
            for j in range(sheet_offset.row - 1, 11, -1):
                sheet[f'{i}{j}'] = None

        class_name = sheet['d3'].value
        wb.save(f'attachments/Template_{class_name}.xlsx')
        wb.close()

    except Exception as create_error:
        return f'Во время создания возникла ошибка ({create_error})'


def clear_template(class_name, date_today):
    x = 13

    wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
    sheet = wb['Лист1']

    columns = ['f', 'g', 'h', 'i', 'j', 'k']

    while sheet[f'b{x}'].value != 'итого':
        x += 1
    sheet_offset = sheet['b13'].offset(row=x - 13, column=0)

    for i in columns:  # очистка ячеек
        for j in range(sheet_offset.row - 1, 11, -1):
            sheet[f'{i}{j}'] = None

    wb.save(f'attachments/Template_{class_name}.xlsx')
    wb.close()

    shutil.copy(f'attachments/Tabel_{class_name}_{date_today}.xlsx',
                f'archive/class_{class_name}/Tabel_{class_name}_{date_today}.xlsx')
    os.remove(f'attachments/Tabel_{class_name}_{date_today}.xlsx')


def send_tabel(class_name=None, mail=None):
    date_today = datetime.datetime.today().date().strftime('%d.%m.%y')

    msg = MIMEMultipart()

    message = f'Табель на {date_today}'
    msg["Subject"] = f'Табель {class_name} по питанию на {date_today}'
    to_email = mail
    msg.attach(MIMEText(message, 'plain'))

    server = smtplib.SMTP('smtp.gmail.com: 587')
    server.starttls()
    try:

        wb = openpyxl.load_workbook(filename=f'attachments/Template_{class_name}.xlsx')
        sheet = wb['Лист1']
        sheet['L9'] = date_today
        month_num = datetime.datetime.now().month
        month_name = calendar.month_name[month_num]
        month_name_russian = {
            'January': 'Январь',
            'February': 'Февраль',
            'March': 'Март',
            'April': 'Апрель',
            'May': 'Май',
            'June': 'Июнь',
            'July': 'Июль',
            'August': 'Август',
            'September': 'Сентябрь',
            'October': 'Октябрь',
            'November': 'Ноябрь',
            'December': 'Декабрь'
        }[month_name]
        sheet['K6'] = month_name_russian

        wb.save(f'attachments/Tabel_{class_name}_{date_today}.xlsx')
        wb.close()

        server.login(from_mail, password)

        file = f'Tabel_{class_name}_{date_today}.xlsx'
        filename = os.path.basename(file)
        ftype, encoding = mimetypes.guess_type(file)
        file_type, subtype = ftype.split('/')

        if file_type == "application":
            with open(f'./attachments/{file}', 'rb') as f:
                file = MIMEApplication(f.read(), subtype)

            file.add_header('content-disposition', 'attachment', filename=filename)
            msg.attach(file)

        server.sendmail(from_mail, to_email, msg.as_string())
        clear_template(class_name, date_today)

        server.close()
        return f'Табель отправлен на почту'

    except Exception as _ex_error:
        return f'Во время отправки произошла ошибка: {_ex_error}'


global add_teacher_name, add_vk_id, add_class_name, add_email

bot = Bot(bot_token)

tprint('LOADED', font='5lineoblique')

conn = sq.connect("teachers.sqlite")
cur = conn.cursor()

confirm_keyboard = {
    Keyboard(inline=True)
    .add(Callback('✅Да', {'cmd_send': 'agree_send', 'cmd_add': 'a'}), KeyboardButtonColor.POSITIVE)
    .add(Callback('❌Нет', {'cmd_send': 'disagree_send', 'cmd_add': 'd'}), KeyboardButtonColor.NEGATIVE)
}

confirm_keyboard_2 = {
    Keyboard(inline=True)
    .add(Callback('✅Да', {'cmd_send': 'a', 'cmd_add': 'agree_add'}), KeyboardButtonColor.POSITIVE)
    .add(Callback('❌Нет', {'cmd_send': 'd', 'cmd_add': 'disagree_add'}), KeyboardButtonColor.NEGATIVE)
}

main_keyboard = {
    Keyboard()
    .add(Text('Выведи табель'), KeyboardButtonColor.POSITIVE)
    .row()
    .add(Text('Список платников'))
    .add(Text('Настройки'))
}

settings_keyboard_true = {
    Keyboard()
    .add(Text('Напоминалка'), KeyboardButtonColor.POSITIVE)
    .add(Text('Назад'), KeyboardButtonColor.NEGATIVE)
}

settings_keyboard_false = {
    Keyboard()
    .add(Text('Напоминалка'), KeyboardButtonColor.NEGATIVE)
    .add(Text('Назад'), KeyboardButtonColor.NEGATIVE)
}


@bot.loop_wrapper.interval(seconds=5)
async def reminder():
    date_today = datetime.datetime.today().date()
    readable_time = int(time.mktime(time.strptime(f"{date_today}  12:00:00", "%Y-%m-%d %H:%M:%S")))
    weekday_today = datetime.date.today().weekday()
    if weekday_today == 6:
        return
    if weekday_today == 5:
        (cur.execute(f'''SELECT vk_id FROM `teachers_table` WHERE reminder = TRUE AND {readable_time + 43200} >= 
                next_answer AND class_name LIKE "1%"'''))
        answer = cur.fetchall()
        for teacher in answer:
            await bot.api.messages.send(peer_id=teacher[0].replace('id', ''),
                                        message="Не забудь написать кого сегодня нет!", random_id=0)
            cur.execute(f'''UPDATE teachers_table SET next_answer = {generate()} WHERE vk_id = "{teacher[0]}"''')
            conn.commit()
    if math.floor(time.time()) >= readable_time:
        (cur.execute(f'''SELECT vk_id FROM `teachers_table` WHERE reminder = TRUE AND {readable_time + 43200} >= 
        next_answer'''))
        answer = cur.fetchall()

        for teacher in answer:
            await bot.api.messages.send(peer_id=teacher[0].replace('id', ''),
                                        message="Не забудь написать кого сегодня нет!", random_id=0)
            cur.execute(f'''UPDATE teachers_table SET next_answer = {generate()} WHERE vk_id = "{teacher[0]}"''')
            conn.commit()


@bot.on.private_message(text='. <missing>')
async def filling_tabel(event: Message, missing):
    payers_dict = {}

    id_user = f"id{(await event.get_user(id)).get('id')}"
    (cur.execute(f'''SELECT class_name FROM `teachers_table` WHERE vk_id == "{id_user}"'''))
    class_n = cur.fetchone()[0].upper()

    missing = missing.split()

    payers = create_list_payers(class_n)
    for payer in payers:
        payers_dict[payer] = 'н'

    for person in payers_dict:
        if person in missing:
            payers_dict[person] = 'н'
        else:
            payers_dict[person] = 1

    cur.execute(f'''UPDATE teachers_table SET next_answer = {generate()} WHERE vk_id = "{id_user}"''')
    conn.commit()

    values_list = list(payers_dict.values())
    filling_template(values_list, class_n)
    await event.answer(f"Сегодня {len(values_list) - values_list.count('н')} платников")


@bot.on.private_message(text='/табель')
async def create_layout(event: Message):
    attach = event.attachments

    if attach.__len__() > 0:
        document = attach[0].doc

        if document is None:
            return

        name, type_f = document.title.split(".")
        if type_f == 'xlsx':
            response = requests.get(document.url)
            with open(document.title, "wb") as f:
                f.write(response.content)
            create_template(document.title)
            await event.answer(type_f)
        elif type_f != 'xlsx':
            await event.answer(
                f'Возможно, отправленный файл не является табелем.\nЕсли ошибка повторяется, напишите @id{admin_id}')
    elif attach.__len__() == 0:
        await event.answer(f'К сообщению не прикреплён файл, прикрепи его, и напиши ту же команду в ОДНОМ сообщении')


@bot.on.private_message(text='@id')
async def id_claimer(event: Message):
    id_person = await event.get_user(id)
    await event.answer(id_person)


@bot.on.private_message(text='Выведи табель')
async def tabel_sender(event: Message):
    await event.answer('Вы подтверждаете отправку табеля?', keyboard=confirm_keyboard)


@bot.on.private_message(text=['/репорт <ticket>', 'репорт <ticket>'])
async def report_handler(message: Message, ticket):
    user_data = await bot.api.users.get(message.from_id)
    await bot.api.messages.send(random_id=0, peer_id=admin_id,
                                message=f'У @id{user_data[0].id}({user_data[0].first_name}'
                                        f' {user_data[0].last_name}) возникла проблема: {ticket}')


@bot.on.private_message(text='/ad <announcement>')
async def announce_handler(message: Message, announcement):
    if message.from_id == admin_id:
        data = []
        cur.execute('''SELECT vk_id FROM teachers_table''')
        for i in cur.fetchall():
            data.append(i[0].replace('id', ''))
        data.remove(str(admin_id))
        await bot.api.messages.send(random_id=0, peer_ids=data, message=announcement)
        await message.answer(message='Объявление было отправлено')


@bot.on.private_message(text=['@add_teacher <add_teach_args>', '@добавить учителя <add_teach_args>'])
async def add_teacher(event: Message, add_teach_args):
    if (await event.get_user(id)).get('id') == admin_id:
        add_teacher_name, add_class_name, add_vk_id, add_email = add_teach_args.split(', ')
        await event.answer(
            message=f'Вы хотите добавить учителя |{add_teacher_name}|\n( @{add_class_name} ) {add_vk_id} ({add_email}) класса. Если все данные верны, подтвердите',
            keyboard=confirm_keyboard_2)


@bot.on.private_message()
async def main_handler(event: Message):
    msg = event.text
    id_user = f"id{(await event.get_user(id)).get('id')}"
    if msg == 'старт' or msg == 'начать':
        await event.answer(message='Бот запущен, вы на главной странице', keyboard=main_keyboard)
    elif msg == 'Список платников':
        cur.execute(f'''SELECT class_name FROM teachers_table WHERE vk_id = "{id_user}"''')
        class_name = cur.fetchone()[0]
        payers = create_list_payers(class_name)
        data = f',\n-'.join(payers)
        await event.answer(message=f'Список платиков в Вашем классе:\n-{data}')
    elif msg == 'Настройки':
        cur.execute(f'''SELECT reminder FROM teachers_table WHERE vk_id = "{id_user}"''')
        remind = cur.fetchone()[0]
        await event.answer(message='Вы на странице настроек бота',
                           keyboard=(settings_keyboard_true if remind == True else settings_keyboard_false))
    elif msg == 'Назад':
        await event.answer(message='Вы на главной странице бота', keyboard=main_keyboard)
    elif msg == 'Напоминалка':
        cur.execute(f'''SELECT reminder FROM teachers_table WHERE vk_id = "{id_user}"''')
        remind = cur.fetchone()[0]
        if remind == True:
            cur.execute(f'''UPDATE teachers_table SET reminder = 0 WHERE vk_id = "{id_user}"''')
            conn.commit()
            await event.answer(message=f'Напоминалка выключена', keyboard=settings_keyboard_false)
        else:
            cur.execute(f'''UPDATE teachers_table SET reminder = 1 WHERE vk_id = "{id_user}"''')
            conn.commit()
            await event.answer(message=f'Напоминалка включена', keyboard=settings_keyboard_true)
    elif msg == 'Помощь':
        await event.answer()


@bot.on.raw_event(GroupEventType.MESSAGE_EVENT, dataclass=GroupTypes.MessageEvent)
async def confirm_handler(event: GroupTypes.MessageEvent):
    cmd_send = event.object.payload['cmd_send']
    cmd_add = event.object.payload['cmd_add']

    user_id = event.object.user_id

    if cmd_send == 'agree_send':
        (cur.execute(f'''SELECT class_name, teacher_email FROM `teachers_table` WHERE vk_id == "id{user_id}"'''))
        class_n, email = cur.fetchone()
        send_tab = send_tabel(class_name=class_n, mail=email)
        await event.ctx_api.messages.edit(peer_id=user_id, conversation_message_id=event.object.conversation_message_id,
                                          message=send_tab)
    elif cmd_send == 'disagree_send':
        await event.ctx_api.messages.edit(peer_id=user_id, conversation_message_id=event.object.conversation_message_id,
                                          message='Табель не отправлен')

    if cmd_add == 'agree_add':
        cur.execute(
            f'''INSERT OR IGNORE INTO teachers_table(teacher_name, vk_id, class_name, teacher_email) VALUES("{add_teacher_name}", "{add_vk_id}", "{add_class_name}", "{add_email}")''')
        conn.commit()
        await event.ctx_api.messages.edit(peer_id=user_id, conversation_message_id=event.object.conversation_message_id,
                                          message='Учитель добавлен в базу данных')
    elif cmd_add == 'disagree_add':
        await event.ctx_api.messages.edit(peer_id=user_id, conversation_message_id=event.object.conversation_message_id,
                                          message='Учитель не добавлен в базу данных')


bot.run_forever()
